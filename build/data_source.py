"""Fuente única de datos del Cuadro de Mando de Fisioterapia AP - DASE Sureste.

Contiene los datos reales Dic-2025 extraídos del proyecto técnico
(Tablas 29-30 para unidades; Tablas 19-22 para el CMI) y genera:
  - data.json   -> datos precargados que se embeben en el HTML
  - Plantilla_Cuadro_Mando_Fisio_DASE.xlsx / .csv -> plantilla estándar
Es la referencia: editar aquí si cambian los datos base.
"""
import json
import csv
import os

PERIODO_BASE = "Dic-2025"

# --- Capa operativa: 19 unidades (Tabla 29 identificación + Tabla 30 indicadores) ---
# Unidad, Codigo, Poblacion, Fisioterapeutas, PctAtendida, DemoraVI, DemoraTto,
# PresionMes, ObjAltaPct, DolorBajaPct, CarteraPct
UNIDADES = [
    ("Perales de Tajuña",   "16010210", 36318, 1, 1.73, 25.61,  4.00, 18.00, 78.20, 78.88, 72.90),
    ("Arganda del Rey",     "16010410", 23673, 1, 2.79, 45.08,  8.93,  9.00, 81.29, 99.12, 97.30),
    ("Santa Mónica",        "16010610", 96175, 4, 2.79, 30.50, 11.09, 13.92, 72.77, 93.54, 50.60),
    ("Ibiza",               "16011410", 64476, 2, 2.21, 19.14,  4.52, 12.95, 88.98, 94.44, 78.40),
    ("Arroyo Medialegua",   "16011710", 47599, 2, 3.04, 10.70,  3.20, 13.64, 85.80, 93.13, 59.20),
    ("Numancia",            "16011810", 65611, 3, 2.87, 36.80,  7.34, 12.02, 87.86, 90.37, 61.80),
    ("Ángela Uriarte",      "16012110", 60925, 3, 3.85, 53.82,  4.29, 16.60, 80.39, 94.95, 50.50),
    ("Vicente Soldevilla",  "16012310", 34056, 2, 3.13, 48.68, 20.68,  9.33, 81.95, 93.44, 98.50),
    ("Buenos Aires",        "16012610", 42539, 2, 3.36, 37.90,  7.94, 12.86, 86.01, 91.48, 82.40),
    ("Federica Montseny",   "16012710", 92374, 4, 3.02, 53.47,  3.90, 16.43, 90.49, 95.05, 84.40),
    ("Villa de Vallecas",   "16012910", 30211, 2, 3.59, 45.90,  2.55, 12.33, 80.04, 74.07, 82.70),
    ("Entrevías",           "16013210", 39378, 2, 3.41, 43.98,  7.46, 13.12, 66.71, 87.22, 77.30),
    ("Pavones",             "16013510", 49372, 2, 2.84, 47.29, 11.94, 11.33, 82.45, 96.97, 75.40),
    ("Villablanca",         "16014010", 80025, 2, 1.82, 79.15, 14.28, 17.95, 80.30, 95.57, 55.40),
    ("Arganda-Felicidad",   "16014110", 55752, 2, 2.26, 62.35,  8.00, 16.24, 44.39, 72.61, 75.40),
    ("Ensanche de Vallecas","16014210", 59020, 2, 2.33, 45.81,  9.33, 13.24, 92.17, 94.82, 48.10),
    ("Valleaguado",         "16021310", 88950, 5, 3.76, 44.13, 10.61, 10.63, 76.89, 96.16, 71.60),
    ("San Fernando",        "16021610", 40483, 2, 2.93, 30.16, 10.06, 14.21, 81.90, 95.69, 58.60),
    ("Mejorada del Campo",  "16021710", 46392, 2, 2.05, 64.17,  5.24, 10.69, 75.29, 95.45, 52.80),
]

UNIDAD_COLS = ["Periodo", "Unidad", "Codigo", "Poblacion", "Fisioterapeutas",
               "PctAtendida", "DemoraVI", "DemoraTto", "PresionMes",
               "ObjAltaPct", "DolorBajaPct", "CarteraPct"]

# --- Capa estratégica: CMI P1-P4 (Tablas 19-22) ---
# Codigo, Indicador, Objetivo, Valor, Unidad, PuntoPartida, Meta(texto),
# Direccion, MetaNum, EstadoManual(None si numérico calculable)
CMI = [
    ("P1-01", "Demora de valoración inicial",            "P1", 43.63, "días", 43.63, "≤ 35 días en 18 meses", "menor_mejor", 35.0, None),
    ("P1-02", "Demora de inicio de tratamiento",         "P1",  8.05, "días",  8.05, "≤ 7 días en 18 meses",  "menor_mejor",  7.0, None),
    ("P1-03", "Presión asistencial y frecuentación",     "P1", 13.40, "ratio",13.40, "Reducir la dispersión entre unidades", "menor_mejor", None, "ambar"),
    ("P1-04", "Actividad grupal (Educación para la Salud)","P1", 0.0,  "% unid.",0.0, "≥ 50% de unidades en 12 meses", "mayor_mejor", 50.0, None),
    ("P2-01", "Consecución de objetivos al alta",        "P2", 79.16, "%",   79.16, "≥ 80% en todas las unidades", "mayor_mejor", 80.0, None),
    ("P2-02", "Mejora del dolor al alta",                "P2", 94.00, "%",   94.00, "≥ 90% en todas las unidades", "mayor_mejor", 90.0, None),
    ("P2-03", "Cumplimiento de cartera (506/414)",       "P2", None,  "%",   None,  "Todas las unidades > 74,5%", "mayor_mejor", 74.5, None),
    ("P2-04", "Población atendida",                      "P2",  2.82, "%",    2.82, "Estable o al alza",     "mayor_mejor", None, "verde"),
    ("P3-01", "Satisfacción durante la lista de espera", "P3", None,  "%",   None,  "Medida y al alza tras el triaje", "mayor_mejor", None, "rojo"),
    ("P3-02", "Triaje en la primera quincena",           "P3",  0.0,  "%",    0.0,  "≥ 80% de derivaciones",  "mayor_mejor", 80.0, None),
    ("P4-01", "Formación en líneas estratégicas",        "P4",  0.0,  "%",    0.0,  "≥ 60% en 24 meses",      "mayor_mejor", 60.0, None),
    ("P4-02", "Coordinación: circuitos de derivación",   "P4",  0.0,  "unid.",0.0,  "Circuito normalizado en las 19 unidades", "mayor_mejor", 19.0, "rojo"),
]

CMI_COLS = ["Periodo", "Codigo", "Indicador", "Objetivo", "Valor", "Unidad",
            "PuntoPartida", "Meta", "Direccion", "MetaNum", "EstadoManual"]

OBJETIVOS = {
    "P1": "Accesibilidad y presión asistencial",
    "P2": "Efectividad clínica y cartera",
    "P3": "Experiencia del paciente y triaje",
    "P4": "Formación y coordinación",
}

# Configuración de los indicadores operativos para el semáforo por celda.
# clave -> (etiqueta, unidad, direccion, metaNum|null, tipo)
# tipo: "meta" (semáforo contra metaNum) o "relativo" (color por desviación de la media)
OPERATIVO_COLS = [
    ("PctAtendida",  "% atendida",      "mayor_mejor", None, "relativo"),
    ("DemoraVI",     "Demora VI (d)",   "menor_mejor", 35.0, "meta"),
    ("DemoraTto",    "Demora tto (d)",  "menor_mejor",  7.0, "meta"),
    ("PresionMes",   "Presión/mes",     "menor_mejor", None, "relativo"),
    ("ObjAltaPct",   "% obj. alta",     "mayor_mejor", 80.0, "meta"),
    ("DolorBajaPct", "% dolor alta",    "mayor_mejor", 90.0, "meta"),
    ("CarteraPct",   "% cartera",       "mayor_mejor", 74.5, "meta"),
    # Actividad (origen: PDF "Informe Unidades de Atención Específica")
    ("PendientesVI", "Pendientes VI",   "menor_mejor", None, "relativo"),
    ("SesionesIndiv","Sesiones indiv.", "mayor_mejor", None, "relativo"),
    ("ConsultasVI",  "Consultas VI",    "mayor_mejor", None, "relativo"),
]


def build_data():
    unidades = []
    for row in UNIDADES:
        rec = dict(zip(UNIDAD_COLS[1:], row))
        rec["Periodo"] = PERIODO_BASE
        unidades.append(rec)

    # media DASE de cartera para P2-03 (estaba "muy dispersa" en el doc)
    media_cartera = round(sum(u["CarteraPct"] for u in unidades) / len(unidades), 2)

    cmi = []
    for row in CMI:
        rec = dict(zip(CMI_COLS[1:], row))
        rec["Periodo"] = PERIODO_BASE
        if rec["Codigo"] == "P2-03":
            rec["Valor"] = media_cartera
            rec["PuntoPartida"] = media_cartera
        cmi.append(rec)

    return {
        "periodos": [PERIODO_BASE],
        "objetivos": OBJETIVOS,
        "operativoCols": [
            {"clave": c[0], "etiqueta": c[1], "direccion": c[2],
             "metaNum": c[3], "tipo": c[4]} for c in OPERATIVO_COLS
        ],
        "unidades": unidades,
        "cmi": cmi,
        "fuente": "Proyecto Técnico Fisioterapia DASE Sureste (Tablas 19-22, 29-30). Línea base Dic-2025.",
    }


def write_json(path):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(build_data(), f, ensure_ascii=False, indent=2)


def _fmt_es(v):
    if v is None:
        return ""
    if isinstance(v, float):
        return ("%.2f" % v).replace(".", ",")
    return v


def write_csv_templates(folder):
    data = build_data()
    # Unidades.csv
    with open(os.path.join(folder, "Plantilla_Unidades.csv"), "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(UNIDAD_COLS)
        for u in data["unidades"]:
            w.writerow([_fmt_es(u[c]) for c in UNIDAD_COLS])
    # CMI.csv
    with open(os.path.join(folder, "Plantilla_CMI.csv"), "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(CMI_COLS)
        for r in data["cmi"]:
            w.writerow([_fmt_es(r[c]) for c in CMI_COLS])


def write_xlsx_template(path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    data = build_data()
    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="1F4E5F")
    hdr_font = Font(bold=True, color="FFFFFF")

    ws = wb.active
    ws.title = "Unidades"
    ws.append(UNIDAD_COLS)
    for u in data["unidades"]:
        ws.append([u[c] for c in UNIDAD_COLS])

    ws2 = wb.create_sheet("CMI")
    ws2.append(CMI_COLS)
    for r in data["cmi"]:
        ws2.append([r[c] for c in CMI_COLS])

    # hoja de instrucciones
    ws3 = wb.create_sheet("LÉEME")
    instr = [
        ["Plantilla del Cuadro de Mando — Fisioterapia AP DASE Sureste"],
        [""],
        ["Cómo usarla:"],
        ["1. Para añadir un periodo nuevo, copia las filas y cambia la columna 'Periodo' (p. ej. Mar-2026)."],
        ["2. No cambies los nombres de las hojas (Unidades, CMI) ni las cabeceras de columnas."],
        ["3. Usa punto o coma decimal indistintamente. Deja la celda vacía si no hay dato (no pongas 0)."],
        ["4. Guarda el archivo y cárgalo en el cuadro de mando (botón 'Cargar datos' o arrastrar y soltar)."],
        [""],
        ["Hoja Unidades — una fila por unidad y periodo:"],
        [", ".join(UNIDAD_COLS)],
        [""],
        ["Hoja CMI — una fila por indicador estratégico y periodo:"],
        [", ".join(CMI_COLS)],
        ["Direccion: 'menor_mejor' (p. ej. demoras) o 'mayor_mejor' (p. ej. % al alta)."],
        ["MetaNum: meta numérica para el semáforo (vacío si la meta es cualitativa)."],
        ["EstadoManual: verde/ambar/rojo para indicadores cualitativos (vacío si se calcula solo)."],
    ]
    for r in instr:
        ws3.append(r)

    for sheet in (ws, ws2):
        for cell in sheet[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
        sheet.freeze_panes = "A2"
        for col in sheet.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            sheet.column_dimensions[col[0].column_letter].width = min(width + 2, 28)
    wb.save(path)


if __name__ == "__main__":
    here = os.path.dirname(os.path.abspath(__file__))
    root = os.path.dirname(here)
    write_json(os.path.join(here, "data.json"))
    plantillas = os.path.join(root, "plantillas")
    os.makedirs(plantillas, exist_ok=True)
    write_csv_templates(plantillas)
    write_xlsx_template(os.path.join(plantillas, "Plantilla_Cuadro_Mando_Fisio_DASE.xlsx"))
    print("OK: data.json + plantillas (.xlsx, .csv) generadas")
